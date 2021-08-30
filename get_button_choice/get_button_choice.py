import openpyxl
from openpyxl import Workbook

def set(path, type, amount, date, userId):
    try:
        data = openpyxl.load_workbook(path)
        # get the first sheet
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])

        table = data.active
        nrows = table.max_row
        ncolumns = table.max_column

        table.cell(nrows + 1, 1).value = str(type)
        table.cell(nrows + 1, 2).value = float(amount)
        table.cell(nrows + 1, 3).value = str(date)
        table.cell(nrows + 1, 4).value = str(userId)

        data.save(path)

    except FileNotFoundError:
        data = Workbook()
        table = data.active
        table['A1']='type'
        table['B1']='amount'
        table['C1']='date'
        table['D1']='userId'

        table.cell(2, 1).value = str(type)
        table.cell(2, 2).value = float(amount)
        table.cell(2, 3).value = str(date)
        table.cell(2, 4).value = str(userId)

        data.save(path)



def edit(path, origin_type, origin_amount, origin_date, type, amount, date, userId):
    data = openpyxl.load_workbook(path)
    # get the first sheet
    sheetnames = data.get_sheet_names()
    table = data.get_sheet_by_name(sheetnames[0])

    table = data.active
    nrows = table.max_row
    ncolumns = table.max_column

    for i in range(1,nrows+1):
        if table.cell(i, 4).value == userId and table.cell(i, 1).value == origin_type and table.cell(i,2).value==origin_amount and table.cell(i,3).value==origin_date:

            table.cell(i, 1).value = type
            table.cell(i, 2).value = amount
            table.cell(i, 3).value = date

    data.save(path)

def delete(path, type, amount, date, userId):
    data = openpyxl.load_workbook(path)
    # get the first sheet
    sheetnames = data.get_sheet_names()
    table = data.get_sheet_by_name(sheetnames[0])

    table = data.active
    nrows = table.max_row
    ncolumns = table.max_column

    for i in range(1, nrows + 1):
        if table.cell(i, 4).value == userId and table.cell(i, 1).value == type and table.cell(i,2).value==amount and table.cell(i,3).value==date:

            table.delete_rows(i)

    data.save(path)

def search(path, userId, type=None, amount=None, date=None):
    list=[]
    data = openpyxl.load_workbook(path)
    # get the first sheet
    sheetnames = data.get_sheet_names()
    table = data.get_sheet_by_name(sheetnames[0])

    table = data.active
    nrows = table.max_row
    ncolumns = table.max_column

    if type==None and amount==None and date==None:
        for i in range(1,nrows+1):
            if str(table.cell(i, 4).value) == userId:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    elif type!=None and amount==None and date==None:
        for i in range(1, nrows + 1):
            if str(table.cell(i, 4).value) == userId and str(table.cell(i, 1).value)==type:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    elif type==None and amount!=None and date==None:
        for i in range(1, nrows + 1):
            if str(table.cell(i, 4).value) == userId and float(table.cell(i, 2).value)==amount:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    elif type==None and amount==None and date!=None:
        for i in range(1, nrows + 1):
            if str(table.cell(i, 4).value) == userId and str(table.cell(i, 3).value)==date:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    elif type!=None and amount!=None and date==None:
        for i in range(1, nrows + 1):
            if str(table.cell(i, 4).value) == userId and str(table.cell(i, 1).value)==type and float(table.cell(i, 2).value)==amount:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    elif type!=None and amount==None and date!=None:
        for i in range(1, nrows + 1):
            if str(table.cell(i, 4).value) == userId and str(table.cell(i, 1).value)==type and str(table.cell(i, 3).value)==date:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    elif type==None and amount!=None and date!=None:
        for i in range(1, nrows + 1):
            if str(table.cell(i, 4).value) == userId and float(table.cell(i, 2).value)==amount and str(table.cell(i, 3).value)==date:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    else:
        for i in range(1, nrows + 1):
            if str(table.cell(i, 4).value) == userId and str(table.cell(i, 1).value)==type and float(table.cell(i, 2).value)==amount and str(table.cell(i, 3).value)==date:
                list.append((str(table.cell(i, 1).value),float(table.cell(i, 2).value),str(table.cell(i, 3).value)))

    return list

#testing
#set('budget.xlsx','food',100,'2020-1-1','ntu001')
#set('budget.xlsx','food',50,'2020-1-20','ntu004')
#set('budget.xlsx','food',150,'2020-3-1','ntu003')
#edit('budget.xlsx','food',100,'2020-1-1','entertainment',200,'2021-1-1','ntu001')
#delete('budget.xlsx','entertainment',200,'2021-1-1','ntu001')
#print(search('budget.xlsx','ntu004'))

